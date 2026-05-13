from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from datetime import datetime, timedelta
import time
import sys
import os

# Set console encoding to UTF-8 to fix printing emojis on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Ensure output and logs directories exist
os.makedirs("output", exist_ok=True)
os.makedirs("logs", exist_ok=True)

# Set up logging to file and stdout
class Logger(object):
    def __init__(self):
        self.terminal = sys.stdout
        log_file = os.path.join("logs", f"scraper_{datetime.now().strftime('%Y%m%d_%H%M')}.log")
        self.log = open(log_file, "a", encoding="utf-8")
        
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
        self.log.flush()
        
    def flush(self):
        self.terminal.flush()
        self.log.flush()

sys.stdout = Logger()

# ==========================================
# CONFIG
# ==========================================
LIMIT_PAGES = 200        # Số trang tối đa (50 bản ghi/trang = 1000 gói)
DAYS_BACK   = 30        # Chỉ lấy gói đăng trong N ngày gần nhất (0 = tất cả)
PAGE_WAIT   = 6         # Giây chờ tối đa sau mỗi lần chuyển trang

KEYWORDS_MOBIFONE = [
    # ── NHÓM 1: Viễn thông & Mạng lưới ──────────────────────────
    "viễn thông", "BTS", "trạm phát sóng", "truyền dẫn", "cáp quang",
    "hạ tầng mạng", "hạ tầng số", "chuyển đổi số", "dịch vụ số",
    "an toàn thông tin", "bảo mật", "cloud", "data center",
    "trung tâm dữ liệu", "internet", "wifi", "sim", "CNTT",
    "hội nghị trực tuyến", "camera giám sát",

    # ── NHÓM 2: Thiết bị mạng & Switch (từ danh mục CTC) ────────
    "switch", "router", "access point", "module SFP", "media converter",
    "ODF", "măng xông", "măng sông", "dây nhảy", "dây nối quang",
    "cáp mạng", "CAT5", "CAT6", "tủ rack", "rack 19",
    "thiết bị mạng", "thiết bị L2", "thiết bị L3",
    "CPE", "SOHO", "VoIP", "E1",

    # ── NHÓM 3: Hạ tầng BTS & Nguồn điện ────────────────────────
    "tủ nguồn", "ắc quy", "acquy", "UPS", "nguồn điện viễn thông",
    "chống sét", "tiếp địa", "máy phát điện", "biến áp",
    "cột bê tông", "cột ăng ten", "cột viễn thông",
    "cáp điện", "tủ outdoor", "tủ indoor",

    # ── NHÓM 4: Cáp & Phụ kiện hạ tầng ──────────────────────────
    "cáp quang", "cáp FO", "cáp treo", "dây thuê bao quang",
    "ống nhựa PVC", "ống HDPE", "ống thép mạ kẽm",
    "băng báo hiệu", "phụ kiện quang", "fast connector",
    "đầu nối quang",

    # ── NHÓM 5: Thiết bị đo kiểm BTS & Quang ────────────────────
    "máy đo OTDR", "OTDR", "máy đo sóng", "máy TEM",
    "đo vùng phủ sóng", "đo nội trở", "bút soi quang",
    "máy thu công suất",

    # ── NHÓM 6: Hệ thống CNTT & Server ──────────────────────────
    "máy chủ", "server", "kiosk", "màn hình ghép", "màn hình LED",
    "hệ thống IOC", "máy quét CCCD", "máy quét mã vạch",
    "phần mềm", "lưu trữ",

    # ── NHÓM 7: Năng lượng mặt trời ──────────────────────────────
    "pin năng lượng mặt trời", "solar", "inverter solar",
    "tấm pin mặt trời",

    # ── NHÓM 8: Điều hòa & Cơ điện hạ tầng ──────────────────────
    "máy điều hoà", "điều hòa không khí", "điều hòa",
]

# ==========================================
# HELPER FUNCTIONS (định nghĩa 1 lần ngoài vòng lặp)
# ==========================================
def get_v(label, text):
    """Tách giá trị theo nhãn trong block text."""
    if label not in text:
        return "N/A"
    return text.split(label)[1].split("\n")[0].replace(":", "").strip()


def parse_date(date_str):
    """Parse ngày từ string dd/mm/yyyy hoặc dd-mm-yyyy."""
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    return None


def is_recent(date_str, days_back):
    """Kiểm tra gói thầu có trong khoảng ngày gần nhất không."""
    if days_back == 0:
        return True
    dt = parse_date(date_str)
    if dt is None:
        return True  # Không parse được → giữ lại để không bỏ sót
    return dt >= datetime.now() - timedelta(days=days_back)


def score_item(ten, linh_vuc):
    """
    Chấm điểm mức độ phù hợp MobiFone + CTC (1–5 sao).
    ⭐⭐⭐⭐⭐ = cực kỳ phù hợp, cả MobiFone lẫn CTC đều có thể tham gia
    ⭐⭐⭐⭐   = phù hợp cao, thuộc thế mạnh cốt lõi
    ⭐⭐⭐     = phù hợp trung bình, có thể cạnh tranh
    ⭐⭐      = phù hợp thấp, cần liên danh
    ⭐        = sơ bộ liên quan, theo dõi thêm
    """
    text = f"{ten} {linh_vuc}".lower()
    score = 1

    # Nhóm cao nhất — MobiFone + CTC đều mạnh
    tier5 = [
        "hạ tầng mạng", "an toàn thông tin", "chuyển đổi số",
        "cáp quang", "truyền dẫn", "BTS", "trạm phát sóng",
        "tủ nguồn viễn thông", "ODF", "switch", "router",
        "thiết bị mạng", "hạ tầng viễn thông",
    ]
    # Nhóm cao — sản phẩm CTC cung cấp trực tiếp
    tier4 = [
        "viễn thông", "module SFP", "media converter", "access point",
        "ắc quy", "acquy", "chống sét", "tủ rack", "máy chủ",
        "server", "cáp mạng", "CAT6", "VoIP", "wifi",
        "màn hình LED", "IOC", "kiosk", "OTDR", "máy đo sóng",
        "điều hòa", "máy phát điện",
    ]
    # Nhóm trung bình — phụ kiện, vật tư
    tier3 = [
        "phần mềm", "cloud", "bảo mật", "lưu trữ", "sim",
        "internet", "camera", "ống nhựa", "ống HDPE",
        "cáp điện", "biến áp", "solar", "pin mặt trời",
        "inverter", "cột bê tông", "fast connector",
    ]
    # Bonus: đơn vị nhà nước (ưu tiên dự thầu)
    gov = ["ubnd", "sở ", "bộ ", "công an", "quân đội",
           "bệnh viện", "trường học", "trường đh", "cục ", "viện "]

    for kw in tier5:
        if kw in text: score += 1.5
    for kw in tier4:
        if kw in text: score += 1.0
    for kw in tier3:
        if kw in text: score += 0.5
    for kw in gov:
        if kw in text: score += 0.5; break

    return min(round(score), 5)


# ==========================================
# SELENIUM SETUP
# ==========================================
def setup_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


# ==========================================
# MAIN
# ==========================================
if __name__ == "__main__":
    driver = setup_driver()
    wait   = WebDriverWait(driver, 40)
    all_data = []

    try:
        url = "https://muasamcong.mpi.gov.vn/web/guest/contractor-selection?render=index"
        print("🚀 Đang truy cập hệ thống...")
        driver.get(url)

        # --- BƯỚC 1: CHỌN 50 BẢN GHI/TRANG ---
        try:
            print("⚙️  Thiết lập hiển thị 50 bản ghi/trang...")
            select_el = wait.until(EC.presence_of_element_located(
                (By.XPATH, "//select[option[@value='50']]")
            ))
            Select(select_el).select_by_value("50")
            # ✅ Dùng WebDriverWait thay sleep cứng
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "content__body__left__item")))
            print("✅ Đã chọn 50 bản ghi.")
        except Exception as e:
            print(f"⚠️  Không chỉnh được số bản ghi: {e}")

        # --- BƯỚC 2: ĐẾM TỔNG SỐ TRANG ---
        try:
            page_nums = driver.find_elements(By.CSS_SELECTOR, "li.number")
            total_found = int(page_nums[-1].text) if page_nums else LIMIT_PAGES
            MAX_PAGE = min(total_found, LIMIT_PAGES)
            print(f"📡 Hệ thống có {total_found} trang. Sẽ quét {MAX_PAGE} trang.")
        except Exception:
            MAX_PAGE = LIMIT_PAGES

        # --- BƯỚC 3: QUÉT DỮ LIỆU ---
        skipped_old = 0
        for page in range(1, MAX_PAGE + 1):
            print(f"📄 Trang {page}/{MAX_PAGE}...", end=" ")

            try:
                # ✅ Chờ element xuất hiện thay vì sleep cứng
                wait.until(EC.presence_of_element_located(
                    (By.CLASS_NAME, "content__body__left__item")
                ))
                time.sleep(1.5)  # nhỏ thôi, chỉ để JS render xong

                items = driver.find_elements(By.CLASS_NAME, "content__body__left__item")
                count_added = 0

                for item in items:
                    try:
                        full_text = item.text

                        # ✅ Lấy link chi tiết gói thầu
                        try:
                            link_el = item.find_element(By.TAG_NAME, "a")
                            link = link_el.get_attribute("href") or "N/A"
                        except Exception:
                            link = "N/A"

                        # ✅ Lấy tên gói thầu
                        try:
                            ten = item.find_element(By.TAG_NAME, "h5").text
                        except Exception:
                            ten = get_v("Tên gói thầu", full_text)

                        ngay_dang   = get_v("Ngày đăng tải thông báo", full_text)
                        linh_vuc    = get_v("Lĩnh vực", full_text)
                        chu_dau_tu  = get_v("Chủ đầu tư", full_text)
                        dia_diem    = get_v("Địa điểm", full_text)
                        ma_tbmt     = get_v("Mã TBMT", full_text)

                        dong_thau   = (
                            get_v("Thời điểm đóng thầu", full_text)
                            if "Thời điểm đóng thầu" in full_text
                            else get_v("Thời điểm bắt đầu chào giá trực tuyến", full_text)
                        )

                        # ✅ Lọc theo ngày nếu bật
                        if DAYS_BACK > 0 and not is_recent(ngay_dang, DAYS_BACK):
                            skipped_old += 1
                            continue

                        all_data.append({
                            "Mã TBMT":          ma_tbmt,
                            "Tên gói thầu":     ten,
                            "Chủ đầu tư":       chu_dau_tu,
                            "Lĩnh vực":         linh_vuc,
                            "Ngày đăng":        ngay_dang,
                            "Đóng thầu":        dong_thau,
                            "Địa điểm":         dia_diem,
                            "Link chi tiết":    link,     # ✅ MỚI
                        })
                        count_added += 1

                    except Exception:
                        continue

                print(f"Thêm {count_added}/{len(items)} gói.")

                # --- CHUYỂN TRANG với retry ---
                if page < MAX_PAGE:
                    for attempt in range(3):
                        try:
                            next_btn = wait.until(EC.element_to_be_clickable(
                                (By.CLASS_NAME, "btn-next")
                            ))
                            driver.execute_script("arguments[0].click();", next_btn)
                            # ✅ Chờ trang mới load thay vì sleep cứng
                            wait.until(EC.staleness_of(items[0]))
                            break
                        except Exception as e:
                            if attempt == 2:
                                print(f"\n⛔ Không chuyển được trang, dừng tại trang {page}.")
                                MAX_PAGE = page
                            time.sleep(PAGE_WAIT)

            except Exception as e:
                print(f"\n⛔ Lỗi trang {page}: {e}")
                break

    finally:
        driver.quit()

        # --- BƯỚC 4: XUẤT FILE ---
        if all_data:
            df_all = pd.DataFrame(all_data).drop_duplicates(subset=["Mã TBMT"])

            # Lọc gói thầu phù hợp MobiFone
            pattern = "|".join(KEYWORDS_MOBIFONE)
            mask = (
                df_all["Tên gói thầu"].str.contains(pattern, case=False, na=False) |
                df_all["Lĩnh vực"].str.contains(pattern, case=False, na=False)
            )
            df_mobi = df_all[mask].copy()

            # ✅ Thêm cột điểm phù hợp
            df_mobi["Điểm phù hợp"] = df_mobi.apply(
                lambda r: "⭐" * score_item(r["Tên gói thầu"], r["Lĩnh vực"]), axis=1
            )

            # ✅ Thêm cột Nhóm hàng CTC (từ danh mục 48 sản phẩm CTC)
            CTC_GROUPS = {
                "Cáp & Phụ kiện quang": [
                    "cáp quang", "cáp fo", "dây thuê bao quang", "dây treo",
                    "măng xông", "măng sông", "ODF", "dây nhảy", "dây nối",
                    "fast connector", "phụ kiện quang", "băng báo hiệu",
                ],
                "Switch & Router & SFP": [
                    "switch", "router", "module sfp", "sfp", "media converter",
                    "access point", "CPE", "VoIP", "bộ chuyển đổi", "E1",
                ],
                "Hạ tầng BTS & Nguồn": [
                    "tủ nguồn", "ắc quy", "acquy", "lithium", "chống sét",
                    "tiếp địa", "máy phát điện", "biến áp", "cột bê tông",
                    "tủ outdoor", "tủ indoor",
                ],
                "Cáp điện & Ống hạ tầng": [
                    "cáp điện", "cáp mạng", "cat5", "cat6",
                    "ống nhựa pvc", "ống hdpe", "ống thép",
                ],
                "Tủ rack & Thiết bị phòng máy": [
                    "tủ rack", "rack 19", "máy chủ", "server",
                    "điều hòa", "điều hoà",
                ],
                "Đo kiểm & Dụng cụ": [
                    "otdr", "máy đo", "bút soi", "tem đo", "đo sóng",
                    "máy thu công suất", "dao cắt", "kìm tuốt",
                ],
                "Màn hình & Kiosk & CNTT": [
                    "màn hình led", "màn hình ghép", "ioc", "kiosk",
                    "máy quét", "cccd", "mã vạch",
                ],
                "Năng lượng mặt trời": [
                    "solar", "pin mặt trời", "tấm pin", "inverter",
                ],
            }

            def classify_ctc(ten):
                t = str(ten).lower()
                matched = []
                for group, kws in CTC_GROUPS.items():
                    if any(k in t for k in kws):
                        matched.append(group)
                return " | ".join(matched) if matched else ""

            df_mobi["Nhóm hàng CTC"] = df_mobi["Tên gói thầu"].apply(classify_ctc)
            df_mobi = df_mobi.sort_values("Điểm phù hợp", ascending=False)

            # Tên file theo ngày
            date_str  = datetime.now().strftime("%Y%m%d_%H%M")
            file_name = os.path.join("output", f"muasamcong_mobifone_{date_str}.xlsx")

            with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
                df_all.to_excel(writer, sheet_name="Tất cả gói thầu", index=False)
                df_mobi.to_excel(writer, sheet_name="Gợi ý cho MobiFone", index=False)

                # ✅ Sheet 3: Chỉ lấy gói CTC có thể cung cấp hàng
                df_ctc = df_mobi[df_mobi["Nhóm hàng CTC"] != ""].copy()
                df_ctc.to_excel(writer, sheet_name="CTC có thể cung cấp", index=False)

                from openpyxl.styles import PatternFill, Font, Alignment
                HEADER_FILLS = {
                    "Tất cả gói thầu":     "1A3E6B",
                    "Gợi ý cho MobiFone":  "1D6B3E",
                    "CTC có thể cung cấp": "7B3F00",
                }

                for sheet_name, color in HEADER_FILLS.items():
                    ws = writer.sheets[sheet_name]
                    # Header màu
                    for cell in ws[1]:
                        cell.fill = PatternFill("solid", fgColor=color)
                        cell.font = Font(color="FFFFFF", bold=True, size=10)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Độ rộng cột
                    ws.column_dimensions["A"].width = 20   # Mã TBMT
                    ws.column_dimensions["B"].width = 55   # Tên gói thầu
                    ws.column_dimensions["C"].width = 35   # Chủ đầu tư
                    ws.column_dimensions["D"].width = 20   # Lĩnh vực
                    ws.column_dimensions["E"].width = 14   # Ngày đăng
                    ws.column_dimensions["F"].width = 20   # Đóng thầu
                    ws.column_dimensions["G"].width = 18   # Địa điểm
                    ws.column_dimensions["H"].width = 18   # Link
                    # Đổ màu xen kẽ cho dễ đọc
                    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        fill_color = "F0F5FF" if i % 2 == 0 else "FFFFFF"
                        for cell in row:
                            if not cell.fill.fgColor.rgb or cell.fill.fgColor.rgb == "00000000":
                                cell.fill = PatternFill("solid", fgColor=fill_color)

                # ✅ Format cột link thành hyperlink
                for sheet_name in ["Tất cả gói thầu", "Gợi ý cho MobiFone", "CTC có thể cung cấp"]:
                    ws = writer.sheets[sheet_name]
                    link_col = None
                    for cell in ws[1]:
                        if cell.value == "Link chi tiết":
                            link_col = cell.column
                            break
                    if link_col:
                        for row in ws.iter_rows(min_row=2, min_col=link_col, max_col=link_col):
                            for cell in row:
                                if cell.value and cell.value != "N/A":
                                    cell.hyperlink = cell.value
                                    cell.value = "🔗 Xem chi tiết"
                                    cell.font = Font(color="0563C1", underline="single")

            print(f"\n{'='*55}")
            print(f"🏁 HOÀN THÀNH!")
            print(f"📊 Tổng gói quét:               {len(df_all)}")
            print(f"📱 Gói tiềm năng MobiFone:      {len(df_mobi)}")
            print(f"🏗️  Gói CTC có thể cung cấp:    {len(df_ctc)}")
            print(f"📅 Bỏ qua gói cũ (>{DAYS_BACK}d):       {skipped_old}")
            print(f"📂 File: {file_name}")
            if len(df_ctc) > 0:
                print(f"\n📦 Phân loại nhóm hàng CTC:")
                for grp, cnt in df_ctc["Nhóm hàng CTC"].str.split(" | ").explode().value_counts().items():
                    print(f"   {grp}: {cnt} gói")
            print(f"{'='*55}")
        else:
            print("❌ Không thu thập được dữ liệu.")
